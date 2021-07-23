# -*- coding: utf-8 -*-

from odoo import fields, models, api
from ldap3 import Server, Connection, SUBTREE, MODIFY_REPLACE, LEVEL
import json
import base64

from openpyxl import Workbook
import os, fnmatch
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

class Organization(models.Model):
    _name = "ad.organizacion"
    _description = "Организации"
    _order = "name"

    name = fields.Char(u'Наименование', required=True)
    active = fields.Boolean('Active', default=True)
   

class Branch(models.Model):
    _name = "ad.branch"
    _description = "Подразделение"
    _order = "name"

    name = fields.Char(u'Наименование 1С', required=True)
    ad_name = fields.Char(u'Наименование в AD')
    adbook_name = fields.Char(u'Наименование в справочнике')
    organization_id = fields.Many2one("ad.organizacion", string="Организация")
    active = fields.Boolean('Active', default=True)
    address = fields.Text(string='Адрес')
    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов')
    is_view_photo_adbook = fields.Boolean(string='Показывать фото в справочнике контактов')
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка", default=10)


class Department(models.Model):
    _name = "ad.department"
    _description = "Управления/отделы"
    _order = "name"

    name = fields.Char(u'Наименование', required=True)
    active = fields.Boolean('Active', default=True)
    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов', default=True)
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка")

class AdGroup(models.Model):
    _name = "ad.group"
    _description = "Группы AD"
    _order = "name"

    name = fields.Char(u'Наименование', required=True)

    distinguished_name = fields.Char(u'AD distinguishedName')
    account_name = fields.Char(u'sAMAccountName')
    object_SID = fields.Char(u'AD objectSID')
    is_ldap = fields.Boolean('LDAP?', default=False)

    active = fields.Boolean('Active', default=True)



flags = [
    [0x0001, 'SCRIPT'],
    [0x0002, 'ACCOUNTDISABLE'],
    [0x0008, 'HOMEDIR_REQUIRED'],
    [0x0010, 'LOCKOUT'],
    [0x0020, 'PASSWD_NOTREQD'],
    [0x0040, 'PASSWD_CANT_CHANGE'],
    [0x0080, 'ENCRYPTED_TEXT_PWD_ALLOWED'],
    [0x0100, 'TEMP_DUPLICATE_ACCOUNT'],
    [0x0200, 'NORMAL_ACCOUNT'],
    [0x0800, 'INTERDOMAIN_TRUST_ACCOUNT'],
    [0x1000, 'WORKSTATION_TRUST_ACCOUNT'],
    [0x2000, 'SERVER_TRUST_ACCOUNT'],
    [0x10000, 'DONT_EXPIRE_PASSWORD'],
    [0x20000, 'MNS_LOGON_ACCOUNT'],
    [0x40000, 'SMARTCARD_REQUIRED'],
    [0x80000, 'TRUSTED_FOR_DELEGATION'],
    [0x100000, 'NOT_DELEGATED'],
    [0x200000, 'USE_DES_KEY_ONLY'],
    [0x400000, 'DONT_REQ_PREAUTH'],
    [0x800000, 'PASSWORD_EXPIRED'],
    [0x1000000, 'TRUSTED_TO_AUTH_FOR_DELEGATION'],
    [0x04000000, 'PARTIAL_SECRETS_ACCOUNT'],
  ]


class Employer(models.Model):
    _name = "ad.employer"
    _description = "Сотрудники"
    _order = "name"

    name = fields.Char(u'ФИО', required=True)
    active = fields.Boolean('Active', default=True)
    is_ldap = fields.Boolean('LDAP?', default=False)

    organization_id = fields.Many2one("ad.organizacion", string="Организация", compute="_compute_organization", store=True)
    branch_id = fields.Many2one("ad.branch", string="Подразделение")
    department_id = fields.Many2one("ad.department", string="Управление/отдел")
    title = fields.Char(u'Должность')

    ip_phone = fields.Char(u'Вн. номер')
    phone = fields.Char(u'Мобильный телефон 1')
    sec_phone = fields.Char(u'Мобильный телефон 2')

    email = fields.Char(u'E-mail')

    username = fields.Char(u'sAMAccountName')
    object_SID = fields.Char(u'AD objectSID')
    distinguished_name = fields.Char(u'AD distinguishedName')
    user_account_control = fields.Char(u'AD userAccountControl')
    user_account_control_result = fields.Char(u'AD userAccountControl result', compute="_get_user_account_control_result")

    photo = fields.Binary('Фото', default=None)

    birthday = fields.Date(string='Дата рождения')


    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов', default=True)
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка")

    is_fired = fields.Boolean(string='Уволен', default=False)
    fired_date = fields.Date(string='Дата увольнения')

    is_vacation = fields.Boolean(string='Отпуск')
    vacation_start_date = fields.Date(string='Дата начала отпуска')
    vacation_end_date = fields.Date(string='Дата окончания отпуска')

    is_btrip = fields.Boolean(string='Командировка')
    btrip_start_date = fields.Date(string='Дата начала командировки')
    btrip_end_date = fields.Date(string='Дата окончания командировки')

    #Блокировки
    is_yaware = fields.Boolean(string='yaware')
    is_usb_block = fields.Boolean(string='Блокировка USB')
    is_mailarchiva = fields.Boolean(string='mailarchiva')
    is_phone_rec = fields.Boolean(string='Запись телефонных разговоров')
    is_socnet_block = fields.Boolean(string='Блокировка соц.сетей')
    is_mesg_block = fields.Boolean(string='Блокировка мессенджеров')
    is_cloud_block = fields.Boolean(string='Блокировка облаков')
    is_email_block = fields.Boolean(string='Блокировка email')
    is_rem_ad_block = fields.Boolean(string='rem_ad_block')
    is_iw = fields.Boolean(string='iw')
    is_backup = fields.Boolean(string='backup')
    is_vpn = fields.Boolean(string='vpn')
    is_vip = fields.Boolean(string='vip')




    @api.depends("branch_id", "branch_id.organization_id")
    def _compute_organization(self):
        for record in self:
            if record.branch_id:
                record.organization_id = record.branch_id.organization_id

    def _get_user_account_control_result(self):
        for record in self:
            if record.user_account_control:
                print("UAC", int(record.user_account_control, 16))
                print("UAC", hex(int(record.user_account_control)))
                print("UAC", int(record.user_account_control))
                result = ''
                for value in flags:
                    if (int(record.user_account_control) | int(value[0])) == int(record.user_account_control):
                        print("+++", value[0])
                        print("+++", int(value[0]))
                        result += value[1] + ','
                record.user_account_control_result = result

    def action_update_from_ldap(self):
        pass


    #Экспорт справочника в Excel
    def _export_adbook(self):
        print("+++ _export_adbook")

        #Стили
        bd = Side(style='thick', color="000000")#Жирные границы ячейки
        bb = Side(style='thin', color="000000")#Стандартные границы ячеек

        #Стиль для заголовков
        def top_cell_style(cell, top=False):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText = True)
            cell.border = Border(left=bb, top=bb, right=bb, bottom=bb)
            cell.fill = PatternFill('solid', fgColor='A8353A') if top else PatternFill('solid', fgColor='92CEFC')
            cell.font = Font(b=True, color='FFFFFF') if top else Font(b=True, color='4D4D4D')

        #Стиль для обычных ячеек
        def basic_cell_style(cell, ip_phone=False):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText = True) if ip_phone else Alignment(horizontal="left", vertical="center", wrapText = True)
            cell.border = Border(left=bb, top=bb, right=bb, bottom=bb)
            cell.fill = PatternFill('solid', fgColor='FFFFFF')
            

        file_name = '/tmp/ETS_Contacts.xlsx'

        wb = Workbook()

        branch_list = self.env['ad.branch'].search([
                                ('active', '=', True),
                                ('is_view_adbook', '=', True),
                            ], order="sequence desc")
        
        for branch in branch_list:

            #Создания листа
            ws = wb.create_sheet(branch.adbook_name  or branch.name )

            # Шапка 
            ws['A1'] = 'ФИО'
            ws.column_dimensions['A'].width = 45
            ws['B1'] = 'Должность'
            ws.column_dimensions['B'].width = 65.10
            ws['C1'] = 'Внутренний номер'
            ws.column_dimensions['C'].width = 24.43
            ws['D1'] = 'Мобильный телефон 1'
            ws.column_dimensions['D'].width = 24.43
            ws['E1'] = 'Мобильный телефон 2'
            ws.column_dimensions['E'].width = 24.43
            ws['F1'] = 'Электронная почта'
            ws.column_dimensions['F'].width = 31.29
            ws.merge_cells('A2:F2')
            for c in 'ABCDEF':
                top_cell_style(ws['%s1' % c], True)
                top_cell_style(ws['%s2' % c])
            ws.row_dimensions[2].height = 75
            ws['A2'] = branch.address or ''
            
            
            #Получаем список управлений/отделов 
            department_list_id = self.env['ad.employer'].read_group([ 
                                                        ('branch_id', '=', branch.id),
                                                        ('active', '=', True),
                                                        ('is_view_adbook', '=', True),
                                                    ], 
                                                        fields=['department_id'], 
                                                        groupby=['department_id']
                                                    )
            #Список id отделв для поиска сотрудников
            department_ids = []
            
            for data in department_list_id:
                d_id, obj = data['department_id']
                department_ids.append(d_id)

            department_list = self.env['ad.department'].search([ 
                                                    ('id', 'in', department_ids),
                                                    ('active', '=', True),
                                                    ('is_view_adbook', '=', True),
                                                ], 
                                                    order="sequence desc"
                                                )
            
            i = 3
            for department in department_list:
                
                ws.merge_cells('A%d:F%d' % (i, i))
                
                for c in 'ABCDEF':
                    top_cell_style(ws['%s%d' % (c, i)])
                    
                ws['A%d' % i] = department.name or ''
                
                i += 1
                
                employer_list = self.env['ad.employer'].search([
                            ('branch_id', '=', branch.id),
                            ('department_id', '=', department.id),
                            ('active', '=', True),
                            ('is_view_adbook', '=', True),
                        ], order="sequence desc")

                for employer in employer_list:
                    
                    ws['A%d' % i] = employer.name or ''
                    basic_cell_style(ws['A%d' % i])
                    ws['B%d' % i] = employer.title or ''
                    basic_cell_style(ws['B%d' % i])
                    ws['C%d' % i] = employer.ip_phone or ''
                    basic_cell_style(ws['C%d' % i], True)
                    ws['D%d' % i] = employer.phone or ''
                    basic_cell_style(ws['D%d' % i])
                    ws['E%d' % i] = employer.sec_phone or ''
                    basic_cell_style(ws['E%d' % i])
                    ws['F%d' % i] = employer.email or ''
                    basic_cell_style(ws['F%d' % i])

                    # if len(str(employer.title)) > 60:
                    #     ws.row_dimensions[i].height = 35
                    

                    i += 1
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            #ws.print_rows = "1:1"
            #ws.print_title_cols = 'A:F'
            #ws.print_title_rows = '1:2'
            #ws.print_title_cols = 'A:F'
            #ws.print_options.horizontalCentered = True
            ws.print_area = 'A1:F%s' % i
                    
        
        
        #Удаляем страницу по умолчанию 
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        
        wb.save(filename = file_name)
        wb.close()

        return file_name

    



    #     #Подключение к серверу AD
    #     LDAP_HOST = self.env['ir.config_parameter'].sudo().get_param('ldap_host')
    #     LDAP_PORT = self.env['ir.config_parameter'].sudo().get_param('ldap_port')
    #     LDAP_USER = self.env['ir.config_parameter'].sudo().get_param('ldap_user')
    #     LDAP_PASS = self.env['ir.config_parameter'].sudo().get_param('ldap_password')
    #     LDAP_SSL = self.env['ir.config_parameter'].sudo().get_param('ldap_ssl')
    #     ldap_search_base = self.env['ir.config_parameter'].sudo().get_param('ldap_search_base')

    #     if LDAP_HOST and LDAP_PORT and LDAP_USER and LDAP_PASS and LDAP_SSL and ldap_search_base:
    #         pass
    #     else:
    #         return False

    #     ldap_server = Server(host=LDAP_HOST, port=int(LDAP_PORT), use_ssl=LDAP_SSL, get_info='ALL')
    #     c = Connection(ldap_server, user=LDAP_USER, password=LDAP_PASS)
    #     c.bind()
    #     filter = '(&(objectClass=person)(sAMAccountName=' + self.username + '))'
    #     res = c.search(search_base=ldap_search_base,
    #                 search_filter=filter,
    #                 search_scope=SUBTREE,
    #                 attributes=['cn','department', 'title', 'ou', 'ipPhone', 'distinguishedName' ])
    #     print("------------------------------------")
    #     print(res)
    #     if res:
    #         emp = c.response[0]
    #         print(emp)
    #         atr = emp['attributes']
    #         dn = emp['dn']
    #         print(atr)
    #         department = atr['department']
    #         self.name = atr['cn']
    #         self.department = atr['department']
    #         self.title = atr['title']
    #         self.ou = dn.split(',OU=')[1]
    #         self.ip_phone = atr['ipPhone']

       
    #     return True





class AdSyncEmployer(models.TransientModel):
    _name = 'ad.sync_employer_wizard'
    _description = "Wizard обновление AD"

    result = fields.Text(string='Результат')

    #@api.model
    def ad_sync_emloyer_action(self):
         #     #Подключение к серверу AD
        LDAP_HOST = self.env['ir.config_parameter'].sudo().get_param('ldap_host')
        LDAP_PORT = self.env['ir.config_parameter'].sudo().get_param('ldap_port')
        LDAP_USER = self.env['ir.config_parameter'].sudo().get_param('ldap_user')
        LDAP_PASS = self.env['ir.config_parameter'].sudo().get_param('ldap_password')
        LDAP_SSL = self.env['ir.config_parameter'].sudo().get_param('ldap_ssl')
        LDAP_SEARCH_BASE = self.env['ir.config_parameter'].sudo().get_param('ldap_search_base')
        LDAP_SEARCH_FILTER = self.env['ir.config_parameter'].sudo().get_param('ldap_search_filter')
        LDAP_SEARCH_GROUP_FILTER = self.env['ir.config_parameter'].sudo().get_param('ldap_search_group_filter')

        if LDAP_HOST and LDAP_PORT and LDAP_USER and LDAP_PASS and LDAP_SSL and LDAP_SEARCH_BASE and LDAP_SEARCH_FILTER and LDAP_SEARCH_GROUP_FILTER:
            pass
        else:
            notification = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': ('Прерванно'),
                    'message': 'Нет учетных данных для подключения',
                    'type':'warning',  #types: success,warning,danger,info
                    'sticky': False,  #True/False will display for few seconds if false
                },
            }
            return notification
        try:
            ldap_server = Server(host=LDAP_HOST, port=int(LDAP_PORT), use_ssl=LDAP_SSL, get_info='ALL', connect_timeout=10)
            c = Connection(ldap_server, user=LDAP_USER, password=LDAP_PASS, auto_bind=True)
        except Exception as e:
            print("ERROR connect AD: ", str(e))
            notification = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': ('Ошибка'),
                    'message': 'Невозможно соединиться с AD. Ошибка: ' + str(e),
                    'type':'warning',  #types: success,warning,danger,info
                    'sticky': False,  #True/False will display for few seconds if false
                },
            }
            return notification

        attributes = ['cn', 'title', 'ipPhone', 'mobile', 'mail', 'department', 'sn', 'memberof', 'distinguishedName', 'homePhone', 'whenChanged', 'objectSID', 'sAMAccountName', 'thumbnailPhoto', 'userAccountControl']
        
        total_entries = 0
        
        res = c.search(
                        search_base=LDAP_SEARCH_BASE,
                        search_filter=LDAP_SEARCH_FILTER,
                        search_scope=SUBTREE,
                        attributes=attributes,
                        paged_size = 500
                    )
        print("------------------------------------")
        total_entries += len(c.response)
        cookie = c.result['controls']['1.2.840.113556.1.4.319']['value']['cookie']
        empl_list = c.entries
        page = 1
        while cookie:
            print("search page = ", page)
            res = c.search(
                            search_base=LDAP_SEARCH_BASE,
                            search_filter=LDAP_SEARCH_FILTER,
                            search_scope=SUBTREE,
                            attributes=attributes,
                            paged_size = 500,
                            paged_cookie = cookie
                        )
            total_entries += len(c.response)
            cookie = c.result['controls']['1.2.840.113556.1.4.319']['value']['cookie']
            empl_list += c.entries

        print("total_entries = ", total_entries)

        n = 0
        message_error = ''
        message_branch = ''
        message_department = ''
        message_empl_update = ''
        message_empl_create = ''
        for empl in empl_list:

            empl_name = empl['cn'].value
            if empl_name == "Перепелкин Сергей Александрович":
                print("+++++++++++++++++++++++++++", empl_name)

            if len(empl_name) == 0:
                message_error += "Не указано CN поля для записи %s, пропускаю \n" % str(empl) 
                break
            n += 1
            #if n>5: break
            #print(empl)
            #Search Branch
            distinguishedName = empl['distinguishedName'].value
            if empl['distinguishedName'].value:
                branch_name = distinguishedName.split(',OU=')[1]
                #print("Branch = ", branch_name)
                branch_search = self.env['ad.branch'].search([('ad_name', '=', branch_name)],limit=1)
                if not branch_search:
                    branch_id = self.env['ad.branch'].create({
                                                'name': branch_name, 
                                                'ad_name': branch_name, 
                                                'adbook_name': branch_name, 
                                            }).id
                    message_branch += branch_name +"\n"
                else:
                    branch_id = branch_search.id
            else:
                message_error += "Для %s не задан branch. \n" % empl_name
                branch_id = None

            #Search Department
            if empl['department'].value:
                department_name = empl['department'].value
                if len(department_name)>0:
                    #print("department_name = ", department_name)
                    department_search = self.env['ad.department'].search([('name', '=', department_name)],limit=1)
                    if not department_search:
                        department_id = self.env['ad.department'].create({
                                                    'name': department_name, 
                                                }).id
                        message_department += department_name +"\n"
                    else:
                        department_id = department_search.id
                else:
                    message_error += "Для %s не задан department. \n" % empl_name
                    department_id = None
            else:
                message_error += "Для %s не задан department. \n" % empl_name
                department_id = None



            #Photo
            if empl['thumbnailPhoto'].value:
                thumbnailPhoto = base64.b64encode(empl['thumbnailPhoto'].value).decode("utf-8")
            else:
                thumbnailPhoto = None

            #Search Employer
            e_search = self.env['ad.employer'].search([
                                        ('object_SID', '=', empl['objectSID']),
                                        '|',
                                        ('active', '=', False), 
                                        ('active', '=', True)
                                    ],limit=1)



            # 514 отключенный пользователь
            uic = int(empl['userAccountControl'].value or '514')
            active = True
            # Если пользователь отключен, ACCOUNTDISABLE	0x0002	2
            if uic | 2 == uic:
                active = False 
            
            vals = {
                    'name': empl_name,
                    'branch_id': branch_id,
                    'department_id': department_id,
                    'title': empl['title'].value,
                    'ip_phone': empl['ipPhone'].value,
                    'phone': empl['mobile'].value,
                    'sec_phone': empl['homePhone'].value,
                    'email': empl['mail'].value,
                    'username': empl['sAMAccountName'].value,
                    'object_SID': empl['objectSID'].value,
                    'distinguished_name': empl['distinguishedName'].value,
                    'user_account_control': empl['userAccountControl'].value,
                    'photo': thumbnailPhoto,
                    'active': active,
                    'is_ldap': True,
                    # 'photo': base64.b64decode(empl['thumbnailPhoto'].value)

                }
            if len(e_search)>0 :
                #print('Обновление ', e_search.name)
                message_empl_update += empl_name + '\n'
                e_search.write(vals)
            else:
                #print('Создание  ', empl_name)
                message_empl_create += empl_name + '\n'
                self.env['ad.employer'].create(vals)
            #print(vals)
        # if res:
        #     emp = c.response[0]
        #     print(emp)
        #     atr = emp['attributes']
        #     dn = emp['dn']
        #     print(atr)
        #     department = atr['department']
            # self.name = atr['cn']
            # self.department = atr['department']
            # self.title = atr['title']
            # self.ou = dn.split(',OU=')[1]
            # self.ip_phone = atr['ipPhone']

        result ='Всего получено из АД %s записей \n' % total_entries
        if not message_error == '':
            result = "\n Обновление прошло с предупреждениями: \n \n" + message_error
        else:
            result = "\n Обновление прошло успешно \n \n"
        if not message_branch == '':
            result = "\n Добавлены филиалы, необходимо назначить имена: \n" + message_branch
        if not message_department == '':
            result = "\n Добавлены управления/отделы: \n" + message_department
        if not message_empl_create == '':
            result += "\n Создны новые сотрудники: \n" + message_empl_create
        if not message_empl_update == '':
            result += "\n Обновлены сотрудники: \n" + message_empl_update
        #print("result sync: ", result)

        self.result = result 

        return {
				'name': 'Message',
				'type': 'ir.actions.act_window',
				'view_type': 'form',
				'view_mode': 'form',
				'res_model': 'ad.sync_employer_wizard',
				'target':'new',
				'context':{
							'default_result':self.result,
							} 
				}
        
    # Синхронизация групп
    def ad_sync_group_action(self):
         #     #Подключение к серверу AD
        LDAP_HOST = self.env['ir.config_parameter'].sudo().get_param('ldap_host')
        LDAP_PORT = self.env['ir.config_parameter'].sudo().get_param('ldap_port')
        LDAP_USER = self.env['ir.config_parameter'].sudo().get_param('ldap_user')
        LDAP_PASS = self.env['ir.config_parameter'].sudo().get_param('ldap_password')
        LDAP_SSL = self.env['ir.config_parameter'].sudo().get_param('ldap_ssl')
        LDAP_SEARCH_BASE = self.env['ir.config_parameter'].sudo().get_param('ldap_search_base')
        LDAP_SEARCH_FILTER = self.env['ir.config_parameter'].sudo().get_param('ldap_search_filter')
        LDAP_SEARCH_GROUP_FILTER = self.env['ir.config_parameter'].sudo().get_param('ldap_search_group_filter')

        if LDAP_HOST and LDAP_PORT and LDAP_USER and LDAP_PASS and LDAP_SSL and LDAP_SEARCH_BASE and LDAP_SEARCH_FILTER and LDAP_SEARCH_GROUP_FILTER:
            pass
        else:
            notification = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': ('Прерванно'),
                    'message': 'Нет учетных данных для подключения',
                    'type':'warning',  #types: success,warning,danger,info
                    'sticky': False,  #True/False will display for few seconds if false
                },
            }
            return notification
        try:
            ldap_server = Server(host=LDAP_HOST, port=int(LDAP_PORT), use_ssl=LDAP_SSL, get_info='ALL', connect_timeout=10)
            c = Connection(ldap_server, user=LDAP_USER, password=LDAP_PASS, auto_bind=True)
        except Exception as e:
            print("ERROR connect AD: ", str(e))
            notification = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': ('Ошибка'),
                    'message': 'Невозможно соединиться с AD. Ошибка: ' + str(e),
                    'type':'warning',  #types: success,warning,danger,info
                    'sticky': False,  #True/False will display for few seconds if false
                },
            }
            return notification

        attributes = ['cn', 'distinguishedName', 'whenChanged', 'objectSID', 'sAMAccountName']
        
        total_entries = 0
        
        res = c.search(
                        search_base=LDAP_SEARCH_BASE,
                        search_filter=LDAP_SEARCH_GROUP_FILTER,
                        search_scope=SUBTREE,
                        attributes=attributes,
                    )
        print("------------------------------------")
        total_entries += len(c.response)
        group_list = c.entries

        print("total_entries = ", total_entries)

        n = 0
        message_error = ''
        message_update = ''
        message_create = ''
        for group in group_list:

            group_name = group['cn'].value

            if len(group_name) == 0:
                message_error += "Не указано CN поля для записи %s, пропускаю \n" % str(group) 
                break

            #Search Group
            g_search = self.env['ad.group'].search([
                                        ('object_SID', '=', group['objectSID']),
                                        '|',
                                        ('active', '=', False), 
                                        ('active', '=', True)
                                    ],limit=1)

            
            vals = {
                    'name': group_name,
                    'account_name': group['sAMAccountName'].value,
                    'object_SID': group['objectSID'].value,
                    'distinguished_name': group['distinguishedName'].value,
                    'active': True,
                    'is_ldap': True,
                }
            if len(g_search)>0 :
                message_update += group_name + '\n'
                g_search.write(vals)
            else:
                message_create += group_name + '\n'
                self.env['ad.group'].create(vals)
        

        result ='Всего получено из АД %s записей \n' % total_entries
        if not message_error == '':
            result = "\n Обновление прошло с предупреждениями: \n \n" + message_error
        else:
            result = "\n Обновление прошло успешно \n \n"
        if not message_create == '':
            result += "\n Создны новые группы: \n" + message_create
        if not message_update == '':
            result += "\n Обновлены группы: \n" + message_update

        self.result = result 

        return {
				'name': 'Message',
				'type': 'ir.actions.act_window',
				'view_type': 'form',
				'view_mode': 'form',
				'res_model': 'ad.sync_employer_wizard',
				'target':'new',
				'context':{
							'default_result':self.result,
							} 
				}
